from typing import TYPE_CHECKING, Optional, Union

if TYPE_CHECKING:
    from .table_class import Table

from .colors import color_to_rgb
from .format_value import NumberFormatSpec


def value_format_to_excel_format(
    value_format: Optional[Union[NumberFormatSpec, dict]],
) -> str:
    """Convert a cell value_format to an Excel number_format.

    Args:
        value_format: Can be a NumberFormatSpec instance, dict, or None

    Returns:
        Excel number format string
    """
    # Handle None
    if value_format is None:
        return "General"

    # Handle dict - convert to NumberFormatSpec
    if isinstance(value_format, dict):
        try:
            value_format = NumberFormatSpec.from_dict(value_format)
        except (KeyError, TypeError, ValueError):
            return "General"

    # Handle NumberFormatSpec instances
    if isinstance(value_format, NumberFormatSpec):
        return _spec_to_excel_format(value_format)

    # Unknown type
    return "General"


def _spec_to_excel_format(spec: NumberFormatSpec) -> str:
    """Convert a NumberFormatSpec to an Excel number format string.

    Args:
        spec: NumberFormatSpec instance

    Returns:
        Excel number format string

    Note:
        Excel has limited support for custom number formats with scales.
        For formats with scale, we fall back to text format and rely on
        the Python-formatted value being written to the cell.
    """
    # If scale is set, we can't represent this in Excel number format
    # The value will already be formatted by format_value(), so use text format
    if spec.scale:
        return "@"  # Text format

    # Build the Excel format string
    # Start with the number format
    if spec.thousands:
        if spec.decimals == 0:
            number_format = "#,##0"
        else:
            number_format = "#,##0." + "0" * spec.decimals
    else:
        if spec.decimals == 0:
            number_format = "0"
        else:
            number_format = "0." + "0" * spec.decimals

    # Add prefix and suffix
    # Excel format: prefix number_format suffix
    # For simple symbols like $ and %, no quotes needed
    # Excel will format them correctly as part of the format string
    if spec.prefix or spec.suffix:
        number_format = f"{spec.prefix}{number_format}{spec.suffix}"

    return number_format


def _import_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        return openpyxl, Alignment, Border, Font, PatternFill, Side, get_column_letter
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl  "
            "(or: pip install pyproforma[excel])"
        ) from e


def _build_workbook(table: "Table"):
    """Build and return an openpyxl Workbook from a Table."""
    openpyxl, Alignment, Border, Font, PatternFill, Side, get_column_letter = _import_openpyxl()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    if not table.cells:
        return workbook

    for row_idx, row in enumerate(table.cells, start=1):
        is_header = row_idx == 1
        for col_idx, cell_data in enumerate(row, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = cell_data.value
            cell.number_format = value_format_to_excel_format(cell_data.value_format)

            font_kwargs = {}
            if is_header or cell_data.bold:
                font_kwargs["bold"] = True
            if cell_data.font_color is not None:
                r, g, b = color_to_rgb(cell_data.font_color)
                font_kwargs["color"] = f"{r:02X}{g:02X}{b:02X}"
            if font_kwargs:
                cell.font = Font(**font_kwargs)

            if cell_data.background_color is not None:
                r, g, b = color_to_rgb(cell_data.background_color)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(
                    start_color=hex_color, end_color=hex_color, fill_type="solid"
                )

            border_kwargs = {}
            if cell_data.bottom_border == "single":
                border_kwargs["bottom"] = Side(style="thin")
            elif cell_data.bottom_border == "double":
                border_kwargs["bottom"] = Side(style="double")
            if cell_data.top_border == "single":
                border_kwargs["top"] = Side(style="thin")
            elif cell_data.top_border == "double":
                border_kwargs["top"] = Side(style="double")
            if border_kwargs:
                cell.border = Border(**border_kwargs)

            if cell_data.align:
                cell.alignment = Alignment(horizontal=cell_data.align)
            elif is_header:
                cell.alignment = Alignment(horizontal="center")

    for col_idx, column in enumerate(worksheet.columns):
        column_letter = get_column_letter(column[0].column)
        if (
            table.col_widths
            and col_idx < len(table.col_widths)
            and table.col_widths[col_idx] is not None
        ):
            worksheet.column_dimensions[column_letter].width = table.col_widths[col_idx] / 7
        else:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    return workbook


def to_excel(table: "Table", filename="table.xlsx"):
    """Export a Table to an Excel file.

    Args:
        table: The Table instance to export.
        filename: Destination file path (default: ``"table.xlsx"``).
    """
    workbook = _build_workbook(table)
    workbook.save(filename)
    workbook.close()


def to_excel_bytes(table: "Table") -> "BytesIO":
    """Export a Table to Excel and return the content as a BytesIO buffer.

    Useful for serving Excel files over HTTP without writing to disk.

    Args:
        table: The Table instance to export.

    Returns:
        BytesIO buffer positioned at the start, ready for reading.
    """
    from io import BytesIO
    buf = BytesIO()
    workbook = _build_workbook(table)
    workbook.save(buf)
    workbook.close()
    buf.seek(0)
    return buf
