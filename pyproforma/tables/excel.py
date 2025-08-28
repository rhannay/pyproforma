from typing import TYPE_CHECKING, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .table_class import Table

from ..constants import ValueFormat

# Color name to Excel hex color mapping
CSS_COLOR_TO_EXCEL = {
    "red": "FF0000",
    "blue": "0000FF",
    "green": "008000",
    "black": "000000",
    "white": "FFFFFF",
    "yellow": "FFFF00",
    "orange": "FFA500",
    "purple": "800080",
    "pink": "FFC0CB",
    "brown": "A52A2A",
    "gray": "808080",
    "grey": "808080",
    "darkblue": "00008B",
    "darkgreen": "006400",
    "darkred": "8B0000",
    "lightblue": "ADD8E6",
    "lightgreen": "90EE90",
    "gold": "FFD700",
    "silver": "C0C0C0",
    "navy": "000080",
    "maroon": "800000",
    "olive": "808000",
    "lime": "00FF00",
    "aqua": "00FFFF",
    "teal": "008080",
    "fuchsia": "FF00FF",
}


def value_format_to_excel_format(value_format: Optional[ValueFormat]) -> str:
    """Convert a cell value_format to an Excel number_format."""
    if value_format is None:
        return "General"  # Excel default
    elif value_format == "no_decimals":
        return "#,##0"  # Number with commas, no decimals
    elif value_format == "two_decimals":
        return "#,##0.00"  # Number with commas and 2 decimals
    elif value_format == "percent":
        return "0%"  # Percentage with no decimals
    elif value_format == "percent_one_decimal":
        return "0.0%"  # Percentage with 1 decimal
    elif value_format == "percent_two_decimals":
        return "0.00%"  # Percentage with 2 decimals
    elif value_format == "percent_two_decinals":  # Handle typo that exists in codebase
        return "0.00%"  # Percentage with 2 decimals (same as correct spelling)
    elif value_format == "str":
        return "@"  # Text format
    elif value_format == "year":
        return "0"  # Integer format (no decimals, no commas)
    else:
        return "General"  # Default fallback


def to_excel(table: "Table", filename="table.xlsx"):
    """Export a Table to Excel with formatting.

    Args:
        table: The Table instance to export
        filename: The Excel filename to create
    """
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write column headers
    for col_idx, column in enumerate(table.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = column.label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal=column.text_align)

    # Write data rows
    for row_idx, row in enumerate(
        table.rows, start=2
    ):  # Start at row 2 (after headers)
        for col_idx, cell_data in enumerate(row.cells, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)

            # Set the value
            cell.value = cell_data.value

            # Apply number formatting based on value_format
            cell.number_format = value_format_to_excel_format(cell_data.value_format)

            # Apply other formatting
            font_kwargs = {}

            # Handle bold
            if cell_data.bold:
                font_kwargs["bold"] = True

            # Handle font color
            if cell_data.font_color is not None:
                color_name = cell_data.font_color.lower().strip()
                if color_name in CSS_COLOR_TO_EXCEL:
                    font_kwargs["color"] = CSS_COLOR_TO_EXCEL[color_name]

            # Apply font if any font properties are set
            if font_kwargs:
                cell.font = Font(**font_kwargs)

            # Handle borders
            border_kwargs = {}
            if cell_data.bottom_border is not None:
                if cell_data.bottom_border == "single":
                    border_kwargs["bottom"] = Side(style="thin")
                elif cell_data.bottom_border == "double":
                    border_kwargs["bottom"] = Side(style="double")

            if cell_data.top_border is not None:
                if cell_data.top_border == "single":
                    border_kwargs["top"] = Side(style="thin")
                elif cell_data.top_border == "double":
                    border_kwargs["top"] = Side(style="double")

            # Apply border if any border properties are set
            if border_kwargs:
                cell.border = Border(**border_kwargs)

            # Set alignment
            if cell_data.align:
                cell.alignment = Alignment(horizontal=cell_data.align)

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass

        # Set column width with some padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    workbook.save(filename)
    print(f"Table exported to {filename}")
